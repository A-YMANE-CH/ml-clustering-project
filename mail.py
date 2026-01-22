import imaplib
import email
from email.header import decode_header
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import requests
import base64
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class ReceiptClassifier:
    def __init__(self, email_address, email_password, landingai_api_key):
        self.email_address = email_address
        self.email_password = email_password
        self.landingai_api_key = landingai_api_key
        self.imap_server = "imap.gmail.com"
        self.smtp_server = "smtp.gmail.com"
        self.smtp_port = 587
        
    def connect_to_email(self):
        """Connect to Gmail using IMAP"""
        try:
            mail = imaplib.IMAP4_SSL(self.imap_server)
            mail.login(self.email_address, self.email_password)
            return mail
        except Exception as e:
            print(f"Failed to connect to email: {e}")
            return None
    
    def get_facture_emails(self, mail):
        """Fetch unread emails with 'facture' in subject"""
        mail.select("inbox")
        status, messages = mail.search(None, 'UNSEEN SUBJECT "facture"')
        email_ids = messages[0].split()
        return email_ids
    
    def extract_attachment(self, msg):
        """Extract image attachment from email"""
        for part in msg.walk():
            if part.get_content_maintype() == 'multipart':
                continue
            
            content_type = part.get_content_type()
            filename = part.get_filename()
            
            if filename or content_type.startswith('image/'):
                if not filename:
                    ext = content_type.split('/')[-1]
                    filename = f"receipt.{ext}"
                
                if (filename.lower().endswith(('.png', '.jpg', '.jpeg', '.pdf', '.gif', '.webp', '.bmp')) 
                    or content_type.startswith('image/')):
                    attachment_data = part.get_payload(decode=True)
                    if attachment_data:
                        return attachment_data, filename
        return None, None
    
    def send_to_landingai(self, image_data, filename):
        """Send receipt image to LandingAI for agentic document extraction"""
        url = "https://api.va.landing.ai/v1/tools/agentic-document-analysis"
        
        headers = {
            "Authorization": f"Basic {self.landingai_api_key}"
        }
        
        is_pdf = filename.lower().endswith('.pdf')
        
        if is_pdf:
            files = [("pdf", (filename, image_data, "application/pdf"))]
        else:
            if filename.lower().endswith('.png'):
                mime_type = "image/png"
            elif filename.lower().endswith(('.jpg', '.jpeg')):
                mime_type = "image/jpeg"
            elif filename.lower().endswith('.webp'):
                mime_type = "image/webp"
            elif filename.lower().endswith('.gif'):
                mime_type = "image/gif"
            else:
                mime_type = "image/png"
            
            files = [("image", (filename, image_data, mime_type))]
        
        data = {
            "include_marginalia": "true",
            "include_metadata_in_markdown": "true"
        }
        
        try:
            print(f"Sending to LandingAI: {url}")
            response = requests.post(url, headers=headers, files=files, data=data, timeout=60)
            response.raise_for_status()
            result = response.json()
            print(f"✓ Successfully received response from LandingAI")
            return result
        except requests.exceptions.HTTPError as e:
            print(f"HTTP Error calling LandingAI: {e}")
            print(f"Status Code: {e.response.status_code}")
            print(f"Response: {e.response.text if e.response else 'No response'}")
            return None
        except requests.exceptions.Timeout:
            print("Request timed out. The document might be too large or the server is busy.")
            return None
        except Exception as e:
            print(f"Error calling LandingAI: {e}")
            return None
    
    def classify_receipt(self, extracted_data):
        """Classify receipt as Hotel or Restaurant"""
        data_str = json.dumps(extracted_data).lower()
        
        hotel_keywords = ['hotel', 'lodging', 'accommodation', 'room', 'night', 'stay', 
                         'check-in', 'check-out', 'resort', 'inn', 'suite', 'motel']
        restaurant_keywords = ['restaurant', 'cafe', 'dine', 'food', 'meal', 'table',
                              'server', 'tip', 'gratuity', 'beverage', 'menu', 'dish',
                              'bistro', 'bar', 'grill', 'kitchen']
        
        hotel_score = sum(1 for keyword in hotel_keywords if keyword in data_str)
        restaurant_score = sum(1 for keyword in restaurant_keywords if keyword in data_str)
        
        if hotel_score > restaurant_score:
            return "Hotel"
        elif restaurant_score > hotel_score:
            return "Restaurant"
        else:
            return "Restaurant"
    
    def create_excel_file(self, extracted_data, classification, original_filename):
        """Create an Excel file with extracted data and classification"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Receipt Data"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        category_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        category_font = Font(bold=True, size=14)
        
        ws['A1'] = "CLASSIFICATION"
        ws['B1'] = classification
        ws['A1'].fill = category_fill
        ws['A1'].font = category_font
        ws['B1'].fill = category_fill
        ws['B1'].font = category_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A2'] = "Original Filename"
        ws['B2'] = original_filename
        ws['A3'] = "Processing Date"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A5'] = "Field"
        ws['B5'] = "Value"
        ws['A5'].fill = header_fill
        ws['B5'].fill = header_fill
        ws['A5'].font = header_font
        ws['B5'].font = header_font
        
        row = 6
        
        def flatten_dict(d, parent_key='', sep='_'):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    for i, item in enumerate(v):
                        if isinstance(item, dict):
                            items.extend(flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
                        else:
                            items.append((f"{new_key}_{i}", str(item)))
                else:
                    items.append((new_key, str(v)))
            return dict(items)
        
        flattened_data = flatten_dict(extracted_data)
        
        for key, value in flattened_data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        ws[f'A{row+1}'] = "Raw JSON Data"
        ws[f'A{row+1}'].font = header_font
        ws[f'A{row+1}'].fill = header_fill
        
        ws[f'A{row+2}'] = json.dumps(extracted_data, indent=2)
        ws.merge_cells(f'A{row+2}:B{row+10}')
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        
        filename = f"receipt_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename
    
    def send_excel_email(self, recipient, classification, excel_filename):
        """Send email with Excel file attachment"""
        msg = MIMEMultipart()
        msg['From'] = self.email_address
        msg['To'] = recipient
        msg['Subject'] = f"Receipt Processing Complete - Classification: {classification}"
        
        body = f"""
Hello,

Your receipt has been processed successfully.

Classification: {classification}

Please find the extracted data in the attached Excel file.

Thank you!
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        try:
            with open(excel_filename, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {os.path.basename(excel_filename)}'
                )
                msg.attach(part)
        except Exception as e:
            print(f"Error attaching Excel file: {e}")
            return
        
        try:
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            server.login(self.email_address, self.email_password)
            server.send_message(msg)
            server.quit()
            print(f"Excel file sent to {recipient}")
        except Exception as e:
            print(f"Failed to send email: {e}")
    
    def process_emails(self):
        """Main process to handle emails"""
        mail = self.connect_to_email()
        if not mail:
            return
        
        email_ids = self.get_facture_emails(mail)
        print(f"Found {len(email_ids)} unread email(s) with 'facture' in subject")
        
        for email_id in email_ids:
            try:
                status, msg_data = mail.fetch(email_id, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])
                
                from_ = msg.get("From")
                subject = msg.get("Subject")
                print(f"\nProcessing email from: {from_}")
                print(f"Subject: {subject}")
                
                attachment_data, filename = self.extract_attachment(msg)
                
                if attachment_data:
                    print(f"Found attachment: {filename}")
                    
                    print("Sending to LandingAI for extraction...")
                    extracted_data = self.send_to_landingai(attachment_data, filename)
                    
                    if extracted_data:
                        classification = self.classify_receipt(extracted_data)
                        print(f"Classification: {classification}")
                        
                        print("Creating Excel file...")
                        excel_filename = self.create_excel_file(extracted_data, classification, filename)
                        print(f"Excel file created: {excel_filename}")
                        
                        sender_email = email.utils.parseaddr(from_)[1]
                        self.send_excel_email(sender_email, classification, excel_filename)
                    else:
                        print("Failed to extract data from receipt")
                else:
                    print("No attachment found in this email")
                    
            except Exception as e:
                print(f"Error processing email: {e}")
        
        mail.close()
        mail.logout()


if __name__ == "__main__":
    EMAIL_ADDRESS = "aymanechkartil22@gmail.com"
    EMAIL_PASSWORD = "qqrz izci phrr orfa"
    LANDINGAI_API_KEY = "pat_iSVzCsVRLuqb8Mw1cyAGASQ7FXvnVyr7"
    
    classifier = ReceiptClassifier(EMAIL_ADDRESS, EMAIL_PASSWORD, LANDINGAI_API_KEY)
    
    print("Starting receipt classification process...")
    print("Looking for emails with 'facture' in subject...")
    classifier.process_emails()
    print("\nProcess completed!")